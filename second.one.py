import csv
import re
import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter

currency_bet = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055}

class DataSet:
    """Класс DataSet содержит методы по обработке данных, собирает статистику по вакансии и выводит результат в консоль
        Attributes:
            file (str): Название файла, может быть, в том числе, и путём, по которому расположен файл
            vacancies (list): Список вакансий, по которой будет собираться отдельная статистика
        """
    def __init__(self, file):
        """Инициализирует объект класса DataSet
                Args:
                    file (str): Название файла
                    vacancies (list): Список вакансий
                """
        self.file = file
        self.vacancies = [Vacancy(vac) for vac in self.csv_filer(*self.csv_reader(file))]

    def delete_html(self, new_html):
        """Функция удаления HTML-тегов и лишних пробелов из поля.
                new_html (str): Очищаемое поле
                >>> DataSet.delete_html("abc")
        'abc'
        >>> DataSet.delete_html("<div>abc</div>")
        'abc'
        >>> DataSet.delete_html("<div>abc")
        'abc'
        >>> DataSet.delete_html("   abc  ")
        'abc'
        >>> DataSet.delete_html(" abc     abd")
        'abc abd'
        >>> DataSet.delete_html(" <div><strong><i>  abc <i>  abd  <string>")
        'abc abd'
        >>> DataSet.delete_html(" <div> abc <iqewqljl> <  div   > abd <i>")
        'abc abd'


        """
        result = re.compile(r'<[^>]+>').sub('', new_html)
        return result if '\n' in new_html else " ".join(result.split())

    def csv_reader(self, file):
        """Считывание csv-файла с проверкой есть ли данные в файле. Возвращает заголовки файла и данные о вакансиях
            file (str): Название считываемого файла.
        """
        reader = csv.reader(open(file, encoding='utf_8_sig'))
        new_vacancies = [row for row in reader]
        if len(new_vacancies) == 0:
            print("Пустой файл")
            exit()
        elif len(new_vacancies[1:]) == 0:
            print("Нет данных")
            exit()
        else:
            return new_vacancies[0], new_vacancies[1:]

    def csv_filer(self, headers, vacancies):
        """Отчищает лист вакансий от пустых элементов, создает словарь вакансий.
           Возвращает лист со словарями для каждой вакансии
           headers (list): Заголовки csv файла
           vacancies (list): Список с писаниями вакансий
        """
        vacancies_list = list(filter(lambda vac: (len(vac) == len(headers) and vac.count('') == 0), vacancies))
        vacanies_dictionary = [dict(zip(headers, map(self.delete_html, vac))) for vac in vacancies_list]
        return vacanies_dictionary

class Vacancy:
    """Класс для представления вакансии
        Attributes:
            name (str): Название вакансии
            salary (int): Среднее значение зарплаты
            area_name (str): Территория, на которой числится вакансия
            published_at (int): Год публикации вакансии
        """
    def __init__(self, dictionary_vac):
        """Инициализирует объект Vacancy, высчитывает среднюю зарплату и производит конвертацию для целочисленных полей
                Args:
                    dictionary_vac: (dict[str]): Словарь вакансии, из которого инициализируются переменные объекта
        """
        self.name = dictionary_vac['name']
        self.salary = Salary(dictionary_vac['salary_from'], dictionary_vac['salary_to'], dictionary_vac['salary_currency'])
        self.area_name = dictionary_vac['area_name']
        self.published_at = dictionary_vac['published_at']

class Salary:
    """Класс для представления зарплаты
            Attributes:
                salary_from (int): Нижняя граница вилки зарплаты
                salary_to (int): Верхняя граница вилки зарплаты
                salary_gross(bool): Оклад указан до вычета налогов
                salary_currency (str): Валюта оклада
           """
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        """Инициализирует объект Salary
        salary_from (int): Нижняя граница вилки зарплаты
        salary_to (int): Верхняя граница вилки зарплаты
        salary_gross(bool): Оклад указан до вычета налогов
        salary_currency (str): Валюта оклада
        >>> Salary(10.0, 20.4, True, 'RUR').salary_from
        10.0
        >>> Salary(10.0, 20.4, True, 'RUR').salary_to
        20.4
        >>> Salary(10.0, 20.4, True, 'RUR').salary_currency
        'RUR'
        >>> Salary(10.0, 20.4, True, 'RUR').salary_gross
        True
                """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency

    def to_rub(self, new_salary: float) -> float:
        """
            Переводит валюту в рубли при помощи словаря currency_bet.
            new_salary (float): Значение оклада
        >>> Salary(10.0, 20, True, 'RUR').to_rub(10.0 + 20)
        30.0
        >>> Salary(10, 20.0, True, 'RUR').to_rub(10 + 20.0)
        30.0
        >>> Salary(10, 20, True, 'EUR').to_rub(10 + 20)
        1797.0
        >>> Salary(10, 20, True, 'AZN').to_rub(10 + 20)
        1070.4
        """
        return new_salary * currency_bet[self.salary_currency]

class Report:
    """Класс отвечает за генерацию excel-файла из собранной статистики
        Attributes:
            years_salary: Динамика уровня зарплат по годам
            years_vacs_count: Динамика количества вакансий по годам
            prof_years_salary: Динамика уровня зарплат по годам для выбранной профессии
            prof_years_vacs_count: Динамика количества вакансий по годам для выбранной профессии
            city_salary_rate: Уровень зарплат по городам (в порядке убывания)
            city_vacs_rate: Доля вакансий по городам (в порядке убывания)
        """
    def __init__(self, years_salary, years_vacs_count, prof_years_salary, prof_years_vacs_count, city_salary_rate,
                 city_vacs_rate):
        """Инициализирует класс Report и создаёт экземпляр класс Workbook, отвечающего за создание excel-таблиц
        years_salary: Динамика уровня зарплат по годам
            """
        self.years_salary = years_salary
        self.years_vacs_count = years_vacs_count
        self.prof_years_salary = prof_years_salary
        self.prof_years_vacs_count = prof_years_vacs_count
        self.city_salary_rate = city_salary_rate
        self.city_vacs_rate = city_vacs_rate

    def generate_excel(self):
        """ Метод создает два листа в файле report.xlsx
                """
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)
        years_sheet = new_workbook.create_sheet('Статистика по годам')
        city_sheet = new_workbook.create_sheet('Статистика по городам')

        def get_style(sheet):
            """ Метод из собранной статистики генерирует файл - report.xlsx,
                где выводится вся собранна статистика в два листа, а также сохраняет файл report.xlsx в директорию,
                откуда запускается этот Python-скрипт
                            """
            for column in sheet.columns:
                new_length = 0
                for col in column:
                    new_side = Side(style="thin", color="000000")
                    col.border = Border(left=new_side, right=new_side, top=new_side, bottom=new_side)
                    if col.value is not None:
                        new_length = max(len(str(col.value)), new_length)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = new_length + 2

        years_sheet_columns = ['Год', 'Средняя зарплата', 'Средняя зарплата - Программист', 'Количество вакансий',
                               'Количество вакансий - Программист']
        for index, column_name in enumerate(years_sheet_columns):
            years_sheet.cell(row=1, column=index + 1, value=column_name).font = Font(bold=True)

        for years in self.years_salary.keys():
            years_sheet.append([years, self.years_salary[years], self.prof_years_salary[years],
                                self.years_vacs_count[years], self.prof_years_vacs_count[years]])
        get_style(years_sheet)

        city_sheet_columns = ['Город', 'Уровень зарплат', ' ', 'Город', 'Доля вакансий']
        for index, column_name in enumerate(city_sheet_columns):
            city_sheet.cell(row=1, column=index + 1, value=column_name).font = Font(bold=True)

        for index, keys in enumerate(self.city_salary_rate.keys()):
            city_sheet.append([keys, self.city_salary_rate[keys], None, list(self.city_vacs_rate.keys())[index],
                               self.city_vacs_rate[list(self.city_vacs_rate.keys())[index]]])
        get_style(city_sheet)

        for cell in city_sheet['E']:
            cell.number_format = FORMAT_PERCENTAGE_00
        new_workbook.save('report.xlsx')

def get_statistic(result, index, new_message, slice=0, reverse=False):
    """ Функция которая выводит данные в виде строк. Например: Динамика уровня зарплат по годам: {2022: 94892}

        result (str): Строка с данными
        new_message (str): Подписанные значение, которое выводится
        """
    slice = len(result) if slice == 0 else slice
    statistic = dict(sorted(result, key = lambda item: item[index], reverse=reverse)[:slice])
    print(f'{new_message}{str(statistic)}')
    return statistic

def get_vacancies_statistic(vacs_list: List[Vacancy], fields, vac_name: str = ''):
    """ Функция который собирает статистику по вакансии

    vacs_list (List[Vacancy]): список всех имеющихся вакансий
    fields(): используемое поле в классе Vacancy
    vac_name(str): Название ваканчии
    """
    statistic_result = {}
    for vac in vacs_list:
        if vac.__getattribute__(fields) not in statistic_result.keys():
            statistic_result[vac.__getattribute__(fields)] = 0
    if vac_name != '':
        vacs_list = list(filter(lambda item: vac_name in item.name, vacs_list))
    for vac in vacs_list:
        statistic_result[vac.__getattribute__(fields)] += 1
    if fields == 'area_name':
        for key in statistic_result.keys():
            statistic_result[key] = round(statistic_result[key] / len(new_data.vacancies), 4)
    return statistic_result

def get_statistic_salary(vacs_list: List[Vacancy], fields, vac_name: str = ''):
    """ Функция который собирает статистику по зарплате

        vacs_list (List[Vacancy]): список всех имеющихся вакансий
        fields(): используемое поле в классе Vacancy
        vac_name(str): Название ваканчии
        """
    statistic_result = {}
    for vac in vacs_list:
        if vac.__getattribute__(fields) not in statistic_result.keys():
            statistic_result[vac.__getattribute__(fields)] = []
    if vac_name != '':
        vacs_list = list(filter(lambda item: vac_name in item.name, vacs_list))
    for vac in vacs_list:
        salary_to = vac.salary.salary_to
        salary_from = vac.salary.salary_from
        statistic_result[vac.__getattribute__(fields)].append(
            vac.salary.to_rub(float(salary_from) + float(salary_to)) / 2)
    for key in statistic_result.keys():
        statistic_result[key] = int(sum(statistic_result[key]) // len(statistic_result[key])) if len(
            statistic_result[key]) != 0 \
            else 0
    return statistic_result

def get_date(date):
    """ Функция для вывода года публикации вакансии в правильном формате.
    date (str): Дата публикации вакансии

    >>> get_date('2022-05-31T17:32:49+0300')
    '31.05.2022'
                    """
    new_date = datetime.datetime.strptime(date, '%Y-%m-%dT%H:%M:%S%z')
    return int(new_date.strftime('%Y'))


file = input("Введите название файла: ")
vacancy = input("Введите название профессии: ")
new_data = DataSet(file)
new_dict = {}
for vacansies in new_data.vacancies:
    vacansies.published_at = get_date(vacansies.published_at)
    if vacansies.area_name not in new_dict.keys():
        new_dict[vacansies.area_name] = 0
    new_dict[vacansies.area_name] += 1

get_objects = list(filter (lambda x: int(len(new_data.vacancies) * 0.01) <= new_dict[x.area_name], new_data.vacancies))
dinamic_salary_level = get_statistic(get_statistic_salary(new_data.vacancies, 'published_at').items(), 0,
                             'Динамика уровня зарплат по годам: ')
dinamic_vac_count = get_statistic(get_vacancies_statistic(new_data.vacancies, 'published_at').items(), 0,
                                 'Динамика количества вакансий по годам: ')
dinamic_years_salary = get_statistic(get_statistic_salary(new_data.vacancies, 'published_at', vacancy).items(), 0,
                                  'Динамика уровня зарплат по годам для выбранной профессии: ')
prof_years_vac_count = get_statistic(get_vacancies_statistic(new_data.vacancies, 'published_at', vacancy).items(), 0,
                                      'Динамика количества вакансий по годам для выбранной профессии: ')
dinamic_city_salary_rate = get_statistic(get_statistic_salary(get_objects, 'area_name').items(), 1,
                            'Уровень зарплат по городам (в порядке убывания): ', 10, True)
dinamic_city_vac_rate = get_statistic(get_vacancies_statistic(get_objects, 'area_name').items(), 1,
                               'Доля вакансий по городам (в порядке убывания): ', 10, True)
Report(dinamic_salary_level, dinamic_vac_count, dinamic_years_salary, prof_years_vac_count, dinamic_city_salary_rate,
       dinamic_city_vac_rate).generate_excel()
